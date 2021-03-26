"""
the awesome:

██╗     ██╗██████╗         █████╗ ███╗   ██╗██████╗ ██████╗ ███████╗
██║     ██║██╔══██╗       ██╔══██╗████╗  ██║██╔══██╗██╔══██╗██╔════╝
██║     ██║██████╔╝       ███████║██╔██╗ ██║██║  ██║██████╔╝█████╗  
██║     ██║██╔══██╗       ██╔══██║██║╚██╗██║██║  ██║██╔══██╗██╔══╝  
███████╗██║██████╔╝██╗    ██║  ██║██║ ╚████║██████╔╝██║  ██║███████╗
╚══════╝╚═╝╚═════╝ ╚═╝    ╚═╝  ╚═╝╚═╝  ╚═══╝╚═════╝ ╚═╝  ╚═╝╚══════╝
"""

"𝙄𝙢𝙥𝙤𝙧𝙩 𝙡𝙞𝙗𝙨"
import os
import re
import datetime
import unidecode
import warnings
from workalendar.america import Brazil
import openpyxl
import calendar
from collections import OrderedDict
import shutil
import glob
import time
import json
import pywinauto
import pytesseract
import pyautogui
import autoit
from PIL import Image, ImageFilter
import pandas as pd
import tkinter
import pyperclip
import json
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

def number_normalizer(pseudNumber):
    """
    Number Normalizer

    Objective:
        1- 10(int) => 10(int); Finalized
        2- 10.00(float) => 10.00(float); Finalized
        3- '10'(str) => 10(int);
        4- '10.00'(str) => 10.00(float);
        5- 'R$ 10'(str) => 10(int);
        6- 'R$ 10.00'(str) => 10.00(float);
        7- True(bool) => 1(int); Finalized
        8- False(bool) => 0(int); Finalized
    return in success:
        clean value
    return in error:
        False
    """

    if isinstance(pseudNumber, bool):
        if pseudNumber:
            return 1
        else:
            return 0
    elif isinstance(pseudNumber, (int, float)):
        return pseudNumber
    elif isinstance(pseudNumber, str):
        negative = '-' in pseudNumber
        if '.' in pseudNumber or ',' in pseudNumber:
            pseudNumber = re.sub('[^0-9.,]', '', pseudNumber)
            if ',' in pseudNumber:
                pseudNumber = pseudNumber.replace(',', '.')
            pseudNumber = re.sub('[.](?=.*[.])', '', pseudNumber)
            pseudNumber = float(pseudNumber)
            if negative:
                return -pseudNumber
            return pseudNumber
        else:
            pseudNumber = re.sub('\D+', '', pseudNumber)
            pseudNumber = int(pseudNumber)
            if negative:
                return -pseudNumber
            return pseudNumber
    else:
        return False


def text_normalizer(*, text, only_normal_digits=False, lower=False, upper=False, only_digits=False, only_numbers=False, remove_separation=False, remove_strip=True, remove_spaces=False):
    """
    Var Resolve

    Objective:
        Normalize a string, without accentuation, without capital letters, without unnecessary spaces and being able to choose between just numbers or just digits
    return in success:
        Clean text
    return in error:
        It is not possible to give an error
    """

    textNormalized = text
    if lower and not upper:
        textNormalized = text.lower()
    if upper and not lower:
        textNormalized = text.upper()
    if only_normal_digits:
        textNormalized = unidecode.unidecode(textNormalized)
    if remove_separation:
        textNormalized = re.sub('[,-./\\\_]+', '', textNormalized)
    if only_digits:
        textNormalized = re.sub('\d+', '', textNormalized)
    if only_numbers:
        textNormalized = re.sub('\D+', '', textNormalized)
    if remove_strip:
        textNormalized = textNormalized.strip()
    if remove_spaces:
        textNormalized = textNormalized.replace(' ', '')
    return textNormalized


def var_resolve(var):
    """
    Var Resolve

    Objective:
        Identify whether a string variable is a number, if it is, return it as a number variable, if not, return it again as a string
    return in success:
        resolved variable
    return in error:
        It is not possible to give an error
    """

    try:
        return float(var)
    except ValueError:
        return str(var)


class Date:
    """
    Data
    """

    def __init__(self, *, date, format="%d/%m/%Y"):
        self.cal = Brazil()
        self.entry_date = date
        self.entry_format = format
        self.format = format
        self.formatted_date = datetime.datetime.strptime(date, format)
        self.week_days = [
            'Segunda-feira',
            'Terça-feira',
            'Quarta-feira',
            'Quinta-Feira',
            'Sexta-feira',
            'Sábado',
            'Domingo'
        ]


    def now(self):
        return datetime.datetime.now().strftime(self.format)


    def add_days(self, days):
        modified_date = self.formatted_date + datetime.timedelta(days=days)
        modified_date = datetime.datetime.strftime(modified_date, self.format)
        self.formatted_date = datetime.datetime.strptime(modified_date, self.format)


    def add_months(self, months):
        modified_date = self.formatted_date + pd.offsets.DateOffset(months=months)
        modified_date = datetime.datetime.strftime(modified_date, self.format)
        self.formatted_date = datetime.datetime.strptime(modified_date, self.format)


    def get_first_work_day_date(self):
        first_month_day = self.get_first_day_date()
        walking_in_month = datetime.datetime.strptime(first_month_day, self.format)
        while not self.cal.is_working_day(walking_in_month):
            walking_in_month = walking_in_month + datetime.timedelta(days=1)
        return datetime.datetime.strftime(walking_in_month, self.format)


    def get_first_day_date(self):
        first_day_date = self.formatted_date.replace(day=1)
        return datetime.datetime.strftime(first_day_date, self.format)


    def get_last_day_date(self):
        year = self.formatted_date.year
        month = self.formatted_date.month
        last_day = calendar.monthrange(year, month)[1]
        last_day_date = self.formatted_date.replace(day=last_day)
        return datetime.datetime.strftime(last_day_date, self.format)


    def is_work_day(self):
        date = datetime.datetime.strftime(self.formatted_date, self.format)
        date = datetime.datetime.strptime(date, self.format)
        return self.cal.is_working_day(date)


    def get_weekday(self):
        weekday = self.formatted_date.weekday()
        return self.week_days[weekday]


    def formatter(self, *, format):
        formatted_date = datetime.datetime.strftime(self.formatted_date, format)
        formatted_date = datetime.datetime.strptime(formatted_date, format)
        self.formatted_date = formatted_date
        self.format = format


    def get_days_between_data(self, data):
        data1 = datetime.datetime.strptime(data, self.format)
        data2 = self.formatted_date
        days = (data1 - data2).days
        return days


    def get(self):
        return {
            'entry_data': self.entry_date,
            'entry_format': self.entry_format,
            'full_data': datetime.datetime.strftime(self.formatted_date, self.format),
            'format': self.format,
            'year': self.formatted_date.year,
            'month': self.formatted_date.month,
            'day': self.formatted_date.day
        }


def date_is_valid(*, date, format='%d/%m/%Y'):
    try:
        datetime.datetime.strptime(date, format)
        return True
    except:
        return False


def get_current_date(*, format='%d/%m/%Y'):
    return datetime.datetime.now().strftime(format)


def filepath_split(*, filepath):
    """
    FilePath Split

    Objective:
        example: C:\\Users\\bpa\\Documents\\AndreLib\\teste\\teste.csv
    return in success:
        basepath: C:\\Users\\bpa\\Documents\\AndreLib\\teste\\
        filename: teste.csv
    return in error:
        raise Exception, because the file is not valid or not is find
    """
    
    if os.path.isfile(filepath):
        return {'basepath': os.path.dirname(filepath), 'filename': os.path.basename(filepath), 'ext': os.path.splitext(filepath)[1]}
    else:
        raise Exception('O arquivo especificado na função não é um arquivo ou não foi encontrado')


def move_file(*, filepath_a, filepath_b):
    """
    Move File

    Objective:
        move file of route a to route b, this process can rename the file, only specify a name in filepath_b
    return in success:
        the new file path
    return in error:
        raise Exception, because the file is not valid or not is find
    """
    
    if os.path.isfile(filepath_a) and (os.path.isfile(filepath_b) or os.path.isdir(filepath_b)):
        shutil.move(filepath_a, filepath_b)
        return filepath_b
    else:
        raise Exception('Ocorreu um erro ao tentar mover o arquivo')


def rename_file(*, filepath, newName):
    """
    Rename File

    Objective:
        rename file to a new name
    return in success:
        the new file name
    return in error:
        raise Exception, because the file is not valid or not is find
    """
    
    if os.path.isfile(filepath):
        filePathSplitted = filepath_split(filepath=filepath)
        basePath = filePathSplitted['basepath']
        extFile = filePathSplitted['ext']
        os.rename(filepath, f'{basePath}\{newName}{extFile}')
        return newName
    else:
        raise Exception('O arquivo especificado na função não é um arquivo ou não foi encontrado')


def remove_file(*, filepath):
    """
    Rename File

    Objective:
        remove file
    return in success:
        file removeded
    return in error:
        raise Exception, because the file is not valid or not is find
    """
    
    if os.path.isfile(filepath):
        os.remove(filepath)
    else:
        raise Exception('O arquivo especificado na função não é um arquivo ou não foi encontrado')


def clean_folder(*, dirpath, ext=False):
    """
    Rename File

    Objective:
        remove files inside selected folder
    return in success:
        files removeded
    return in error:
        raise Exception, because the path is not valid or not is find
    """
    
    if os.path.isdir(dirpath):
        if ext:
            for file in glob.glob(f"{dirpath}\*.{ext}"):
                os.remove(file)
        else:
            for file in glob.glob(f"{dirpath}\*.*"):
                os.remove(file)
    else:
        raise Exception('O caminho especificado na função não é válido ou não foi encontrado')


def loop_find_file(*, dirpath, ext=False, timeout=False):
    """
    Loop Find File

    Objective:
        performs an infinite loop until a file with the specified extension is found in the specified folder
    return in success:
        return found file
    return in error:
        raise Exception, because the path is not valid or not is find
    """
 
    if os.path.isdir(dirpath):
        start_time = datetime.datetime.now()
        while True:
            if (datetime.datetime.now() - start_time).total_seconds() >= timeout and timeout:
                raise Exception('O arquivo não foi encontrado e o timeout ativou')
            if ext:
                for file in glob.glob(f"{dirpath}\*.{ext}"):
                    return file
            else:
                for file in glob.glob(f"{dirpath}\*.*"):
                    return file
    else:
        raise Exception('O caminho especificado na função não é válido ou não foi encontrado')


def get_json_from_csv(
    *,
    csv,
    separator=';',
    automatic_header=True,
    custom_headers=[],
    create_numbering_column=False,
    remove_first_row=False,
    remove_last_row=False,
    remove_first_item=False,
    remove_last_item=False
):
    """
    Get Json from CSV

    Objective:
        transform csv in json
    return in success:
        return json builded
    return in error:
        there is no way to return error
    """

    csv_builded = []
    csv = csv.split('\n') # Break csv in lines

    if remove_first_row: # Remove first line
        csv = csv[1:]
    if remove_last_row: # Remove last line
        csv = csv[:-1]

    for idxRow, row in enumerate(csv):
        newRow = {}
        
        csv_splitted = row.split(separator)

        if remove_first_item: # Remove first item
            csv_splitted = csv_splitted[1:]
        if remove_last_item: # Remove last item
            csv_splitted = csv_splitted[:-1]

        for idxItem, item in enumerate(csv_splitted):
            if automatic_header and idxRow == 0:
                custom_headers.append(item) # Add custom header
            else:
                try:
                    title = custom_headers[idxItem] # Try set custom header
                except IndexError:
                    title = idxItem # in error case, set header with idx
                newRow[title] = item
        
        # not create a new row with automatic_header, only keep for use as header
        if not (automatic_header and idxRow == 0):
            csv_builded.append(newRow)

    # Creating a numbering column
    if create_numbering_column:
        for idxRow, row in enumerate(csv_builded):
            row = OrderedDict(row) # Create a ordered dict
            row.update({'N°':idxRow+1}) # Insere a new Item (N°) and the row number
            row.move_to_end('N°', last=False) # Insere in first position
            csv_builded[idxRow] = dict(row) # Converte to dict

    # Return
    return csv_builded


def get_xlsx_from_json(*, json, xls_name='default', mark_column_name=False, mark_column_number=False, special_columns=False):
    """
    Get XLS From Json

    Objective:
        transform json in xlsx
    return in success:
        return name file as success
    return in error:
        raise Exception, because the path is not valid or not is find
    """

    wb = openpyxl.Workbook() # Create a new workbook
    ws = wb.active # Call the Worksheet of the workbook

    # Creating a special column
    if special_columns:
        if type(special_columns) == list:
            for idxRow, row in enumerate(json):
                for special_column in special_columns:
                    row[special_column] = ''
                    json[idxRow] = row # add
        elif type(special_columns) == str:
            for idxRow, row in enumerate(json):
                row[special_columns] = ''
                json[idxRow] = row # add
    
    for i in range(len(json)):
        sub_obj = json[i]
        if i == 0 :
            keys = list(sub_obj.keys())
            for k in range(len(keys)):
                # Add the Items title
                row_number = i + 1 # Row number
                column_number = k + 1 # Column number
                ws.row_dimensions[row_number].height = 25
                title = ws.cell(row = row_number, column = column_number)
                title.value = keys[k] # Set the title
                title.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center') # add alignment style

                # If this is a special column for apply different style
                if (mark_column_name and (str(keys[k]) == str(mark_column_name))) or (mark_column_number and k == mark_column_number):
                    title.font = openpyxl.styles.Font(bold=True, size=13, color='000000') # Apply font style
                    title.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color('E13C3C')) # Apply fill color
                else:
                    title.font = openpyxl.styles.Font(bold=True, size=12, color='000000') # Apply font style
                    title.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color('B1B1B1')) # Apply fill color
                
                # Insere border
                title.border = openpyxl.styles.borders.Border(
                    left=openpyxl.styles.borders.Side(style='thin'), 
                    right=openpyxl.styles.borders.Side(style='thin'), 
                    top=openpyxl.styles.borders.Side(style='thin'), 
                    bottom=openpyxl.styles.borders.Side(style='thin')
                )
        for j in range(len(keys)):
            # Add the Items
            row_number = i + 2 # Row number
            column_number = j + 1 # Column number
            ws.row_dimensions[row_number].height = 25
            item = ws.cell(row = row_number, column = column_number)
            try:
                item.value = sub_obj[str(keys[j])]
            except KeyError:
                item.value = ''
            item.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # If this is a special column for apply different style
            if (mark_column_name and (str(keys[j]) == str(mark_column_name))) or (mark_column_number and j == mark_column_number):
                item.font = openpyxl.styles.Font(bold=False, size=11, color='000000') # Apply font style
                item.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color('E5B7B6')) # Apply fill color
            else:
                item.font = openpyxl.styles.Font(bold=False, size=10, color='000000') # Apply font style
                item.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color('E9E7E6')) # Apply fill color
            
            # Insere border
            item.border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style='dashDot'), 
                right=openpyxl.styles.borders.Side(style='dashDot'), 
                top=openpyxl.styles.borders.Side(style='thin'), 
                bottom=openpyxl.styles.borders.Side(style='thin')
            )
    
    # Adding automatic column width
    for column_cells in ws.columns:
        gap_size = 10
        length = max(len(str(cell.value)) for cell in column_cells)
        
        ws.column_dimensions[column_cells[0].column_letter].width = length + gap_size
    # Save xlsx
    wb.save(f'{xls_name}.xlsx')

    # Return name of the xlsx
    return xls_name


def timestamp():
    """
    Timestamp

    Objective:
        generate a serial number with the year, month, day, hour, minute, seconds and the milliseconds
    return in success:
        return the timestamp
    """

    currentTimestamp = re.sub('\D+', '',str(datetime.datetime.now()))
    return currentTimestamp


def send_keys(*, keys, repeat=1, acceptSpecialChars=False):
    if acceptSpecialChars:
        keys = re.sub('[+^%~()]', r'{\g<0>}', keys)
        keys = keys.replace(' ', '{SPACE}')
    for i in range(repeat):
        pywinauto.keyboard.send_keys(keys)
    return keys
  

class Clipboard:
    """
    Clipboard
    """

    @staticmethod
    def get():
        return pyperclip.paste()


    @staticmethod
    def clear():
        tkinter.Tk().clipboard_clear()


    @staticmethod
    def define(text):
        pyperclip.copy(text)


    @staticmethod
    def copy():
        pywinauto.keyboard.send_keys('^c')


warnings.simplefilter('ignore', category=UserWarning)
class AutomaticGui:
    """
    AutomaticGui
    """

    def __init__(self):
        self.app = pywinauto.Application() # This is only visual
        self.win32_app = pywinauto.Application() # This is only visual
        self.allow_exception = False


    def start_app(self, app):
        pywinauto.Application().start(app)


    def connect_in_app(self, **kwargs):
        return self.ConnectInApp(automaticGui_props=self, kwargs=kwargs)
    class ConnectInApp:
        def __init__(self, *, automaticGui_props, kwargs):
            self.automaticGui_props = automaticGui_props
            self.connect_app_parameters = kwargs
        def run(self):
            try:
                app = pywinauto.Application('uia').connect(**self.connect_app_parameters)
                win32_app = pywinauto.Application('win32').connect(**self.connect_app_parameters)
                self.automaticGui_props.app = app
                self.automaticGui_props.win32_app = win32_app
            except pywinauto.findwindows.ElementAmbiguousError:
                app = pywinauto.Application('uia').connect(**self.connect_app_parameters, found_index = 0)
                win32_app = pywinauto.Application('win32').connect(**self.connect_app_parameters, found_index = 0)
                self.automaticGui_props.app = app
                self.automaticGui_props.win32_app = win32_app
        def wait_ready(self, *, timeout=False):
            self.automaticGui_props.wait_window(window=lambda: self.run(), timeout=timeout)
        def wait_first(self, *, timeout=False):
            keep_connect_app_parameters = self.connect_app_parameters
            start_time = datetime.datetime.now()
            while True:
                if (datetime.datetime.now() - start_time).total_seconds() >= timeout and timeout:
                    raise Exception('Nenhuma janela foi encontrada e timeout ativou')
                for idx, connect_app_parameter in enumerate(keep_connect_app_parameters.items()):
                    key = text_normalizer(text=connect_app_parameter[0], only_digits=True)
                    value = connect_app_parameter[1]
                    connect_app_parameter_builded = { key: value }
                    self.connect_app_parameters = connect_app_parameter_builded
                    try:
                        self.run()
                        return idx
                    except:
                        pass


    def app_exists(self, *, timeout=False, **kwargs):
        if timeout != False:
            try:
                self.wait_window(window=lambda: pywinauto.Application('uia').connect(**kwargs), timeout=timeout)
                return True
            except:
                return False
        else:
            try:
                pywinauto.Application('uia').connect(**kwargs)
                return True
            except:
                return False


    def close_app_if_exists(self, **kwargs):
        app = AutomaticGui()
        if app.app_exists(**kwargs):
            app.connect_in_app(**kwargs).run()
            app.kill()


    def window(self, **kwargs):
        return self.Window(automaticGui_props=self, kwargs=kwargs)
    class Window:
        def __init__(self, *, automaticGui_props, kwargs):
            self.automaticGui_props = automaticGui_props
            self.connect_app_parameters = kwargs
            self.click = self.Click(automaticGui_props=automaticGui_props, kwargs=kwargs)
            self.get_text = self.GetText(automaticGui_props=automaticGui_props, kwargs=kwargs)
        class Click:
            def __init__(self, *, automaticGui_props, kwargs):
                self.automaticGui_props = automaticGui_props
                self.connect_app_parameters = kwargs
            def run(self, *, x=5, y=5):
                self.automaticGui_props.app.top_window().window(**self.connect_app_parameters).click_input(coords=(x, y))
            def wait_ready(self, *, timeout=False, x=5, y=5):
                self.automaticGui_props.wait_window(window=lambda: self.run(x=x, y=y), timeout=timeout)
        class GetText:
            def __init__(self, *, automaticGui_props, kwargs):
                self.automaticGui_props = automaticGui_props
                self.connect_app_parameters = kwargs
            def run(self):
                return self.automaticGui_props.app.top_window().window(**self.connect_app_parameters).window_text()
            def wait_ready(self, *, timeout=False):
                return self.automaticGui_props.wait_window(window=lambda: self.run(), timeout=timeout)
        def exists(self, *, timeout=False):
            start_time = datetime.datetime.now()
            if timeout:
                while True:
                    if (datetime.datetime.now() - start_time).total_seconds() >= timeout and timeout:
                        try:
                            return self.automaticGui_props.app.top_window().window(**self.connect_app_parameters).exists()
                        except:
                            return False
                    try:
                        result = self.automaticGui_props.app.top_window().window(**self.connect_app_parameters).exists()
                        if result:
                            return True
                    except:
                        pass
            else:
                try:
                    return self.automaticGui_props.app.top_window().window(**self.connect_app_parameters).exists()
                except:
                    return False
        def wait_close(self, *, timeout=31536000):
            try:
                return self.automaticGui_props.app.top_window().window(**self.connect_app_parameters).wait_not('exists', timeout=timeout, retry_interval=0.1)
            except:
                return True
        def define_as_window(self):
            return self.automaticGui_props.app.top_window().window(**self.connect_app_parameters)
        def get_children(self, **kwargs):
            return self.automaticGui_props.app.top_window().window(**self.connect_app_parameters).children(**kwargs)


    def image(self, *, image_path, confidence=0.9, in_app=False):
        return self.Image(automaticGui_props=self, image_path=image_path, confidence=confidence, in_app=in_app)
    class Image:
        def __init__(self, *, automaticGui_props, image_path, confidence, in_app):
            self.automaticGui_props = automaticGui_props
            self.image_path = image_path
            self.confidence = confidence
            self.in_app = in_app
            self.click = self.Click(automaticGui_props=automaticGui_props, image_path=image_path, confidence=confidence, in_app=in_app)
            self.get_text = self.GetText(automaticGui_props=automaticGui_props, image_path=image_path, confidence=confidence, in_app=in_app)

        class Click:
            def __init__(self, *, automaticGui_props, image_path, confidence, in_app):
                self.automaticGui_props = automaticGui_props
                self.image_path = image_path
                self.confidence = confidence
                self.in_app = in_app
            def run(self, *, x=0, y=0):
                imageCoords = self.automaticGui_props.get_image_position(image_path=self.image_path, confidence=self.confidence, in_app=self.in_app)
                if imageCoords != None:
                    px, py = imageCoords
                    px = px+x
                    py = py+y
                    pyautogui.click(px, py)
                return imageCoords
            def wait_ready(self, *, timeout=False, x=0, y=0):
                imageCoords = self.automaticGui_props.wait_image(image=lambda: self.automaticGui_props.get_image_position(image_path=self.image_path, confidence=self.confidence, in_app=self.in_app), timeout=timeout)
                px, py = imageCoords
                px = px+x
                py = py+y
                pyautogui.click(px, py)
                return imageCoords
        class GetText:
            def __init__(self, *, automaticGui_props, image_path, confidence, in_app):
                self.automaticGui_props = automaticGui_props
                self.image_path = image_path
                self.confidence = confidence
                self.in_app = in_app
            def run(self, *, x=0, y=0, h=0, w=0, optimize=False, image_save_path=False, only_chars=False):
                imageCoords = self.automaticGui_props.get_image_position(image_path=self.image_path, confidence=self.confidence, in_app=self.in_app, with_size=True)
                if imageCoords != None:
                    left, top, width, height = imageCoords
                    left += x
                    top += y
                    width += w
                    height += h
                    image = pyautogui.screenshot(region=(left, top, width, height))
                    if optimize:
                        image = image.resize((round(width*2), round(height*3)), resample=Image.AFFINE)
                        image = image.filter(ImageFilter.GaussianBlur(1))
                    else:
                        image = image.resize((round(width*2), round(height*3)), resample=Image.AFFINE)
                        image = image.filter(ImageFilter.GaussianBlur(1))
                    config = '--psm 6'
                    if only_chars:
                        config += f' -c tessedit_char_whitelist={only_chars}'
                    texto = pytesseract.image_to_string(image, config=config)
                    texto = text_normalizer(text=texto)
                    if image_save_path:
                        image.save(image_save_path)
                    return texto
                return imageCoords
            def wait_ready(self, *, timeout=False, x=0, y=0, h=0, w=0, optimize=False, image_save_path=False, only_chars=False):
                imageCoords = self.automaticGui_props.wait_image(image=lambda: self.automaticGui_props.get_image_position(image_path=self.image_path, confidence=self.confidence, in_app=self.in_app, with_size=True), timeout=timeout)
                left, top, width, height = imageCoords
                left += x
                top += y
                width += w
                height += h
                image = pyautogui.screenshot(region=(left, top, width, height))
                if optimize:
                    image = image.resize((round(width*2), round(height*3)), resample=Image.AFFINE)
                    image = image.filter(ImageFilter.GaussianBlur(1))
                else:
                    image = image.resize((round(width*2), round(height*2)), resample=Image.AFFINE)
                    image = image.filter(ImageFilter.GaussianBlur(1))
                config = '--psm 6'
                if only_chars:
                    config += f' -c tessedit_char_whitelist={only_chars}'
                texto = pytesseract.image_to_string(image, config=config)
                texto = text_normalizer(text=texto)
                if image_save_path:
                    image.save(image_save_path)
                return texto
        def exists(self, timeout=False):
            start_time = datetime.datetime.now()
            if timeout:
                while True:
                    imageCoords = self.automaticGui_props.get_image_position(image_path=self.image_path, confidence=self.confidence, in_app=self.in_app)
                    if (datetime.datetime.now() - start_time).total_seconds() >= timeout and timeout:
                        if imageCoords == None:
                            return False
                        else:
                            return True
                    elif imageCoords != None:
                        return True
            else:
                imageCoords = self.automaticGui_props.get_image_position(image_path=self.image_path, confidence=self.confidence, in_app=self.in_app)
                if imageCoords == None:
                    return False
                else:
                    return True
        def wait_close(self, *, timeout=False):
            start_time = datetime.datetime.now()
            while True:
                imageCoords = self.automaticGui_props.get_image_position(image_path=self.image_path, confidence=self.confidence, in_app=self.in_app)
                if (datetime.datetime.now() - start_time).total_seconds() >= timeout and timeout:
                    if imageCoords == None:
                        return True
                    else:
                        raise Exception('Imagem não foi fechada e timeout ativou')
                elif imageCoords == None:
                    return True

    def wait_window(self, *, window, timeout):
        waitWindow = True
        start_time = datetime.datetime.now()
        while waitWindow:
            if (datetime.datetime.now() - start_time).total_seconds() >= timeout and timeout:
                raise pywinauto.findwindows.ElementNotFoundError
            try:
                returnWindow = window()
            except pywinauto.findwindows.ElementNotFoundError:
                waitWindow = True
            except pywinauto.findbestmatch.MatchError:
                waitWindow = True
            else:
                return returnWindow


    def wait_window_close(self, *, timeout):
        try:
            self.app.top_window().wait_not('exists', timeout=timeout)
        except:
            pass


    def get_image_position(self, *, image_path, confidence, in_app, with_size=False):
        if not with_size:
            if in_app:
                appCoords = self.get_coords()
                print(appCoords)
                imageCoords = pyautogui.locateCenterOnScreen(image_path, confidence=confidence, region=(appCoords.left, appCoords.top, appCoords.right-appCoords.left, appCoords.bottom-appCoords.top))
            else:
                imageCoords = pyautogui.locateCenterOnScreen(image_path, confidence=confidence)
        else:
            if in_app:
                appCoords = self.get_coords()
                imageCoords = pyautogui.locateOnScreen(image_path, confidence=confidence, grayscale=True, region=(appCoords.left, appCoords.top, appCoords.right-appCoords.left, appCoords.bottom-appCoords.top))
            else:
                imageCoords = pyautogui.locateOnScreen(image_path, confidence=confidence)
        return imageCoords


    def wait_image(self, *, image, timeout):
        waitImage = True
        start_time = datetime.datetime.now()
        while waitImage:
            if (datetime.datetime.now() - start_time).total_seconds() >= timeout and timeout:
                raise Exception('Imagem não foi encontrada e timeout ativou')
            returnImage = image()
            if returnImage != None:
                return returnImage
            else:
                waitImage = True


    def exists(self):
        try:
            return self.app.top_window().exists()
        except:
            return False


    def focus(self):
        self.app.top_window().set_focus()


    def maximize(self):
        self.app.top_window().maximize()


    def restore_size(self):
        self.app.top_window().restore()


    def minimize(self):
        self.app.top_window().minimize()


    def get_coords(self):
        return self.app.top_window().rectangle()


    def get_title(self):
        return self.app.top_window().element_info.name


    def get_process_id(self):
        return self.app.top_window().element_info.process_id


    def printIdentifiers(self):
        print(self.app.top_window().print_control_identifiers())
    
    
    def close(self):
        self.app.top_window().close()


    def kill(self):
        os.system(f'taskkill /F /PID {self.get_process_id()}')


class AutomaticWeb:
    """
    AutomaticWeb
    """

    def __init__(self, *, chromeWebdriver, visible=True, setDefaultPathDownload=False):
        chrome_options = webdriver.ChromeOptions()
        prefs = {
            'safebrowsing.enabled': True,
            'download.directory_upgrade': True
        }

        if setDefaultPathDownload:
            prefs['download.default_directory'] = setDefaultPathDownload
        chrome_options.add_experimental_option('prefs', prefs)
        if not visible:
            chrome_options.add_argument("--headless")
        self.browser = webdriver.Chrome(executable_path=chromeWebdriver, chrome_options=chrome_options)


    def start_website(self, link_website):
        self.browser.get(link_website)
        self.browser.maximize_window()
    

    def element(self, **kwargs):
        return self.Element(automaticWeb_props=self, kwargs=kwargs)
    class Element:
        def __init__(self, *, automaticWeb_props, kwargs):
            self.automaticWeb_props = automaticWeb_props
            search_per_and_value = list(kwargs.items())
            connect_element_parameters_search_per = search_per_and_value[0][0]
            connect_element_parameters_value = search_per_and_value[0][1]
            connect_element_parameters = (eval(f'By.{connect_element_parameters_search_per}'), connect_element_parameters_value)
            self.connect_element_parameters = connect_element_parameters
            self.click = self.Click(automaticWeb_props=automaticWeb_props, connect_element_parameters=connect_element_parameters)
            self.send_keys = self.SendKeys(automaticWeb_props=automaticWeb_props, connect_element_parameters=connect_element_parameters)
        class Click:
            def __init__(self, *, automaticWeb_props, connect_element_parameters):
                self.automaticWeb_props = automaticWeb_props
                self.connect_element_parameters = connect_element_parameters
            def run(self):
                self.automaticWeb_props.browser.find_element(self.connect_element_parameters[0], self.connect_element_parameters[1]).location_once_scrolled_into_view
                self.automaticWeb_props.browser.find_element(self.connect_element_parameters[0], self.connect_element_parameters[1]).click()
            def wait_ready(self, *, timeout=False):
                WebDriverWait(self.automaticWeb_props.browser, timeout).until(EC.element_to_be_clickable(self.connect_element_parameters)).location_once_scrolled_into_view
                WebDriverWait(self.automaticWeb_props.browser, timeout).until(EC.element_to_be_clickable(self.connect_element_parameters)).click()
        class SendKeys:
            def __init__(self, *, automaticWeb_props, connect_element_parameters):
                self.automaticWeb_props = automaticWeb_props
                self.connect_element_parameters = connect_element_parameters
            def run(self, *, keys):
                self.automaticWeb_props.browser.find_element(self.connect_element_parameters[0], self.connect_element_parameters[1]).location_once_scrolled_into_view
                self.automaticWeb_props.browser.find_element(self.connect_element_parameters[0], self.connect_element_parameters[1]).send_keys(keys)
            def wait_ready(self, *, keys, timeout=False):
                WebDriverWait(self.automaticWeb_props.browser, timeout).until(EC.presence_of_element_located(self.connect_element_parameters)).location_once_scrolled_into_view
                WebDriverWait(self.automaticWeb_props.browser, timeout).until(EC.presence_of_element_located(self.connect_element_parameters)).send_keys(keys)


    def get_title(self):
        return self.browser.title


    def close(self):
        self.browser.quit()

